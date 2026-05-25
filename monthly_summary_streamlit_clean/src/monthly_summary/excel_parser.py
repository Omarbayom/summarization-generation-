from __future__ import annotations

import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def canon_note(text: Any) -> str:
    if text is None:
        return ""
    t = str(text).strip().lower()
    t = t.replace("•", " ").replace("–", " ").replace("—", " ").replace("-", " ")
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"[,:;]+", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def parse_sheet_date(sheet_name: str, default_year: int) -> Optional[datetime]:
    s = str(sheet_name).strip().lower()
    s = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", s)
    m = re.search(r"\b(\d{1,2})\s+([a-z]+)(?:\s+(\d{4}))?\b", s)
    if not m:
        return None
    day = int(m.group(1))
    month_word = m.group(2)
    year = int(m.group(3)) if m.group(3) else int(default_year)
    month = MONTHS.get(month_word)
    if month is None:
        for key, value in MONTHS.items():
            if key.startswith(month_word):
                month = value
                break
    if month is None:
        return None
    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def merged_bounds_for_cell(ws, row: int, col: int) -> Optional[Tuple[int, int, int, int]]:
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col
    return None


def cell_value(ws, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return norm(val)
    bounds = merged_bounds_for_cell(ws, row, col)
    if bounds:
        min_row, _, min_col, _ = bounds
        return norm(ws.cell(row=min_row, column=min_col).value)
    return ""


def find_first_cell(ws, text: str, lookat_part: bool = True) -> Optional[Tuple[int, int]]:
    target = str(text).strip().lower()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = cell_value(ws, r, c).strip().lower()
            if not v:
                continue
            if (lookat_part and target in v) or (not lookat_part and target == v):
                return r, c
    return None


def find_marker_row(ws, text: str) -> Optional[int]:
    found = find_first_cell(ws, text, lookat_part=True)
    return found[0] if found else None


def build_skip_rows(ws, excel_cfg: Dict[str, Any]) -> set[int]:
    skip: set[int] = set()
    start_row = find_marker_row(ws, excel_cfg["start_marker"])
    if start_row is not None:
        end_row = None
        for marker in excel_cfg.get("end_markers", []):
            found = find_marker_row(ws, marker)
            if found is not None:
                end_row = found
                break
        if end_row is not None and end_row >= start_row:
            skip.update(range(start_row, end_row + 1))

    blockers_row = find_marker_row(ws, excel_cfg["blockers_marker"])
    support_row = find_marker_row(ws, excel_cfg["support_marker"])

    if blockers_row is not None:
        if support_row is not None and support_row > blockers_row:
            skip.update(range(blockers_row, support_row))
        else:
            skip.update(range(blockers_row, min(ws.max_row, blockers_row + 15) + 1))

    if support_row is not None:
        clear_rows = int(excel_cfg.get("support_clear_rows", 25))
        skip.update(range(support_row, min(ws.max_row, support_row + clear_rows) + 1))

    return skip


def find_header_row(ws, excel_cfg: Dict[str, Any]) -> Optional[int]:
    category_header = str(excel_cfg["category_header"]).strip().lower()
    max_scan = min(ws.max_row, int(excel_cfg.get("header_scan_rows", 100)))
    for r in range(1, max_scan + 1):
        count = 0
        for c in range(1, ws.max_column + 1):
            if cell_value(ws, r, c).strip().lower() == category_header:
                count += 1
        if count >= 2:
            return r
    return None


def extract_pairs_from_two_rows(ws, header_row: int, excel_cfg: Dict[str, Any]) -> List[Tuple[int, int, str]]:
    category_header = str(excel_cfg["category_header"]).strip().lower()
    special_subteams = {"mechanical", "electrical", "control", "system"}
    pairs: List[Tuple[int, int, str]] = []
    seen = set()

    for c in range(2, ws.max_column + 1):
        left = cell_value(ws, header_row, c - 1).lower()
        team = cell_value(ws, header_row, c)
        if team and left == category_header:
            if team.strip().lower() == "system":
                continue
            key = (c - 1, c, team.strip())
            if key not in seen:
                pairs.append(key)
                seen.add(key)

    sub_row = header_row + 1
    if sub_row <= ws.max_row:
        for c in range(2, ws.max_column + 1):
            left = cell_value(ws, sub_row, c - 1).lower()
            sub = cell_value(ws, sub_row, c)
            if sub and left == category_header and sub.strip().lower() in special_subteams:
                team_name = f"System ({sub.strip()})"
                key = (c - 1, c, team_name)
                if key not in seen:
                    pairs.append(key)
                    seen.add(key)
    return pairs


def find_section_ranges(ws) -> Tuple[Optional[Tuple[int, int]], Optional[Tuple[int, int]]]:
    y_cell = find_first_cell(ws, "Yesterday", lookat_part=True)
    t_cell = find_first_cell(ws, "Today", lookat_part=True)

    def marker_range(cell: Optional[Tuple[int, int]]) -> Optional[Tuple[int, int]]:
        if not cell:
            return None
        r, c = cell
        bounds = merged_bounds_for_cell(ws, r, c)
        if bounds:
            min_row, max_row, _, _ = bounds
            return min_row, max_row
        return r, r

    y_range = marker_range(y_cell)
    t_range = marker_range(t_cell)

    if y_cell and t_cell and y_range and t_range and y_range[0] == y_range[1] and t_range[0] == t_range[1]:
        yr = y_cell[0]
        tr = t_cell[0]
        if yr < tr:
            y_range = (yr + 1, tr - 1)
            t_range = (tr + 1, ws.max_row)
    return y_range, t_range


def bucket_by_row(row_idx: int, y_range: Optional[Tuple[int, int]], t_range: Optional[Tuple[int, int]]) -> str:
    if y_range and y_range[0] <= row_idx <= y_range[1]:
        return "Progress"
    if t_range and t_range[0] <= row_idx <= t_range[1]:
        return "Plan"
    return "Other"


def row_empty_for_pairs(ws, row_idx: int, pairs: List[Tuple[int, int, str]]) -> bool:
    for cat_c, team_c, _ in pairs:
        if cell_value(ws, row_idx, cat_c) or cell_value(ws, row_idx, team_c):
            return False
    return True


def extract_items_from_xlsx(input_xlsx: str, cfg: Dict[str, Any]) -> pd.DataFrame:
    excel_cfg = cfg["excel"]
    wb = load_workbook(input_xlsx, data_only=True)
    records = []

    for ws in wb.worksheets:
        header_row = find_header_row(ws, excel_cfg)
        if header_row is None:
            continue
        pairs = extract_pairs_from_two_rows(ws, header_row, excel_cfg)
        if not pairs:
            continue

        skip_rows = build_skip_rows(ws, excel_cfg)
        y_range, t_range = find_section_ranges(ws)
        sheet_dt = parse_sheet_date(ws.title, int(cfg["project"].get("default_year", 2026)))

        row = header_row + 2
        empty_streak = 0
        empty_stop = int(excel_cfg.get("empty_streak_stop", 10))
        while row <= ws.max_row and empty_streak < empty_stop:
            if row in skip_rows:
                row += 1
                continue
            if row_empty_for_pairs(ws, row, pairs):
                empty_streak += 1
                row += 1
                continue
            empty_streak = 0
            bucket = bucket_by_row(row, y_range, t_range)

            for category_col, note_col, team_name in pairs:
                category = cell_value(ws, row, category_col)
                note = cell_value(ws, row, note_col)
                if category.strip().lower() == str(excel_cfg["category_header"]).lower():
                    continue
                if category or note:
                    records.append(
                        {
                            "Sheet": ws.title,
                            "Date": sheet_dt,
                            "Team": team_name,
                            "Bucket": bucket,
                            "Category": category if category else "NULL",
                            "Note": note,
                        }
                    )
            row += 1

    df = pd.DataFrame(records, columns=["Sheet", "Date", "Team", "Bucket", "Category", "Note"])
    if not df.empty:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df.sort_values(["Date", "Team", "Bucket", "Category"], ascending=[True, True, True, True]).reset_index(drop=True)
    return df


def aggregate_items_for_llm(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df.empty:
        return pd.DataFrame(columns=["Team", "Bucket", "Category", "FirstDate", "LastDate", "Items", "Count"])

    data = raw_df.copy()
    data["Date"] = pd.to_datetime(data["Date"], errors="coerce")
    data["Team"] = data["Team"].fillna("").astype(str).str.strip()
    data["Bucket"] = data["Bucket"].fillna("Other").astype(str).str.strip()
    data["Category"] = data["Category"].fillna("NULL").replace("", "NULL")
    data["Note"] = data["Note"].fillna("").astype(str).str.strip()
    data = data[(data["Team"] != "") & (data["Note"] != "")]
    data["Canon"] = data["Note"].apply(canon_note)
    data = data.drop_duplicates(subset=["Team", "Bucket", "Category", "Canon"], keep="first")
    data = data.sort_values(["Team", "Bucket", "Category", "Date"], ascending=[True, True, True, True])

    def to_bullets(series: pd.Series) -> str:
        return "\n".join(f"- {x}" for x in series.tolist() if str(x).strip())

    grouped = (
        data.groupby(["Team", "Bucket", "Category"], as_index=False)
        .agg(
            FirstDate=("Date", "min"),
            LastDate=("Date", "max"),
            Items=("Note", to_bullets),
            Count=("Note", "count"),
        )
        .sort_values(["Team", "Bucket", "Category"], ascending=[True, True, True])
        .reset_index(drop=True)
    )
    return grouped


def build_team_input(agg_df: pd.DataFrame, team: str, max_chars: int) -> str:
    team_rows = agg_df[agg_df["Team"] == team].copy()
    blocks = []
    for _, row in team_rows.iterrows():
        blocks.append(
            f"Bucket: {row['Bucket']}\n"
            f"Category: {row['Category']}\n"
            f"Dates: {row.get('FirstDate', '')} to {row.get('LastDate', '')}\n"
            f"Items:\n{row['Items']}"
        )
    text = "\n\n---\n\n".join(blocks).strip()
    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n[TRUNCATED: input was too long; summarize from the available content only.]"
    return text
